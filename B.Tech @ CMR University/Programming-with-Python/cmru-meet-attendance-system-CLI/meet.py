"""
Copyright (C) 2020 Kritik Agarwal - All Rights Reserved
Unauthorized copying of this file, via any medium is strictly prohibited
Written by Kritik Agarwal, 2020
Support For Windows 10
"""
# import required modules
from math import ceil
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from plyer import notification
from time import sleep
from configparser import ConfigParser
from os import system, name
from datetime import date
import math
import stdiomask
import openpyxl


config_file = 'config.ini'
config = ConfigParser()
config.read(config_file)


def notify(msg):
    message = msg
    notification.notify(title='Attention!', message=message,
                        app_icon='icon.ico', timeout=20, toast=False)


def watermark():
    system('cls')
    print('-'*135)
    print('\033[92m',  '          ██████╗███╗   ███╗██████╗  ', '\033[96m',
          '  ██╗   ██╗███╗   ██╗██╗██╗   ██╗███████╗██████╗ ███████╗██╗████████╗██╗   ██╗ ')
    print('\033[92m',  '         ██╔════╝████╗ ████║██╔══██╗ ', '\033[96m',
          '  ██║   ██║████╗  ██║██║██║   ██║██╔════╝██╔══██╗██╔════╝██║╚══██╔══╝╚██╗ ██╔╝ ')
    print('\033[92m',  '         ██║     ██╔████╔██║██████╔╝ ', '\033[96m',
          '  ██║   ██║██╔██╗ ██║██║██║   ██║█████╗  ██████╔╝███████╗██║   ██║    ╚████╔╝  ')
    print('\033[92m',  '         ██║     ██║╚██╔╝██║██╔══██╗ ', '\033[96m',
          '  ██║   ██║██║╚██╗██║██║╚██╗ ██╔╝██╔══╝  ██╔══██╗╚════██║██║   ██║     ╚██╔╝   ')
    print('\033[92m',  '         ╚██████╗██║ ╚═╝ ██║██║  ██║ ', '\033[96m',
          '  ╚██████╔╝██║ ╚████║██║ ╚████╔╝ ███████╗██║  ██║███████║██║   ██║      ██║    ')
    print('\033[92m',  '          ╚═════╝╚═╝     ╚═╝╚═╝  ╚═╝ ', '\033[96m',
          '   ╚═════╝ ╚═╝  ╚═══╝╚═╝  ╚═══╝  ╚══════╝╚═╝  ╚═╝╚══════╝╚═╝   ╚═╝      ╚═╝    ', '\033[0m')
    print('-'*135)
    print('\033[36m', '\033[1m', '''       __  __  __  __     __       __ _____      ______ __     __          __ __          ___ __         ___  __     
        / _ /  \/  \/ _ |  |_   |\/||_ |_  |    /\  |  | |_ |\ ||  \ /\ |\ |/  |_    /\ /  \ | /  \|\/| /\  | |/  \|\ |
        \__)\__/\__/\__)|__|__  |  ||__|__ |   /--\ |  | |__| \||__//--\| \|\__|__  /--\\\\__/ | \__/|  |/--\ | |\__/| \|
    ''', '\033[0m')
    print('-'*135)
    print('\033[92m', '\033[1m',
          '\n\t\t\t\t\t\tDeveloped By ~ Kritik Agarwal\n', '\033[0m')
    print('-'*135)
    return


def mainmenu():
    watermark()
    print('\nMain Menu')
    print('1) Take Attendance')
    print('2) Setup Email/Password & Excel(.xlsx file)')
    print('3) Set Meet Links & Excel Sheet')
    print('4) Exit')
    try:
        choice = int(input('Select Your Choice: '))
        if choice == 1:
            attendance()
        elif choice == 2:
            epass()
        elif choice == 3:
            meetsetup()
        elif choice == 4:
            exit()
        else:
            print('Wrong Choice')
    except ValueError:
        print('Wrong Choice')
    return


def epass():
    watermark()
    while True:
        mail = input(
            'Enter Your Official "@cmr.edu.in" or "@cmr.ac.in" mail id: ')
        check = mail[mail.index('@') + 1:]
        if check == 'cmr.edu.in' or check == 'cmr.ac.in':
            print('\033[1m', '\033[93m',
                  '\n>>> For Security Concerns, When You Type in the Password it will be masked by asterisk sign ( * ). ', '\033[0m')
            password = stdiomask.getpass(
                '\nEnter Your Password (cAsE sEnSiTiVe): ')
            excelf = input(
                '\nEnter you excel file name with .xlsx in the extension which contains student names (cAsE sEnSiTiVe): ')
            config.set('account', 'email', mail)
            config.set('account', 'pass', password)
            config.set('account', 'excelfile', excelf)
            with open(config_file, 'w') as configfile:
                config.write(configfile)
            print(
                '\033[92m', '\n>>> Successfully saved the Email, Password & Excel File Name.', '\033[0m')
            break
        else:
            print('\033[91m', '>>> Invalid Email.', '\033[0m')
            continue
    input('\n>>> Press any key to return to Main Menu')
    mainmenu()


def meetsetup():
    watermark()
    while ans := 1:
        s = input(
            'Enter a short name for your meet link & excel sheet (for instance 3A or 3B or 5A or 5B): ')
        link = input('Copy Paste The Permanent Meet Link: ')
        excel_sheet = input(
            'Enter the excel sheet name as it is (cAsE sEnSiTiVe): ')  # 3A-CS
        config.set('meeturl', s, link)
        config.set('excelsheet', s, excel_sheet)
        config.set('column', s, 'B')
        with open(config_file, 'w') as configfile:
            config.write(configfile)
        print(
            f'\n>>> Config Saved for {s} as Sheet: {excel_sheet} & Meet Link: {link} ')
        print('-'*50)
        print('\nMenu')
        print('1) Add Another Class Link \n2) Return to Main Menu')
        ans = int(input('Select Your Choice: '))
    input('>>> Press any key to return to Main Menu')
    mainmenu()


def attendance():
    watermark()
    nickname = config['meeturl']
    print(">>> Available Classes (Nickname): ")
    for i in nickname:
        print("---> ", i.upper())
    classroom = input('\nEnter Class Nickname To Take Attendance: ')
    excel_f = config['account']['excelfile']
    excel_sh = config['excelsheet'][classroom]
    gmail_id = config['account']['email']
    gmail_password = config['account']['pass']
    meet_link = config['meeturl'][classroom]
    opt = Options()
    opt.add_argument("--disable-infobars")
    opt.add_argument("start-maximized")
    opt.add_argument("--disable-extensions")
    opt.add_experimental_option("prefs", {
        "profile.default_content_setting_values.media_stream_mic": 1,
        "profile.default_content_setting_values.media_stream_camera": 1,
        "profile.default_content_setting_values.geolocation": 1,
        "profile.default_content_setting_values.notifications": 1
    })
    driver = webdriver.Chrome(
        options=opt, executable_path=r"C:\chromedriver\chromedriver.exe")
    driver.get('https://accounts.google.com/signin/v2/identifier?ltmpl=meet&continue=https%3A%2F%2Fmeet.google.com%3Fhs%3D193&_ga=2.159644724.887115149.1602468384-1763864472.1602468384&flowName=GlifWebSignIn&flowEntry=ServiceLogin%27')
    sleep(2)
    driver.find_element_by_xpath('//input[@type="email"]').send_keys(gmail_id)
    driver.find_element_by_xpath('//*[@id="identifierNext"]').click()
    sleep(3)
    driver.find_element_by_xpath(
        '//input[@type="password"]').send_keys(gmail_password)
    driver.find_element_by_xpath('//*[@id="passwordNext"]').click()
    sleep(4)
    driver.get(meet_link)
    sleep(5)
    cam_mic_selectors = driver.find_elements_by_css_selector(
        'div.U26fgb.JRY2Pb.mUbCce.kpROve')
    for e in cam_mic_selectors:
        e.click()
    sleep(2)
    driver.find_element_by_css_selector(
        'div.uArJ5e.UQuaGc.Y5sE8d.uyXBBb.xKiqt').click()  # join now
    sleep(6)
    notify('Attendance Process Started!\nKindly have some patience till it completes.')
    driver.find_element_by_css_selector(
        'div.uArJ5e.UQuaGc.kCyAyd.QU4Gid.foXzLb.IeuGXd').click()  # participant list
    sleep(6)
    n = int(driver.find_element_by_css_selector('div.eUyZxf span.rua5Nb').text.strip(
        '(').strip(')'))  # no. of participants present
    css_selector = ".tmIkuc"  # For Scrolling
    scrolling_step = 100  # In pixels
    # height (in pixels) for scrolling
    height = driver.execute_script(
        "return document.querySelector('" + css_selector + "').scrollHeight")
    times = math.ceil(height / scrolling_step)
    for i in range(times):
        names = driver.find_elements_by_css_selector(
            'div.KsCJ0 div.kvLJWc span.ZjFb7c')  # participants
        sleep(1)
        driver.execute_script("document.querySelector('" + css_selector +
                              "').scrollTop = " + str(scrolling_step * (i + 1)))  # Scrolling
        sleep(0.25)
    print("\n\n>>> People's found:")
    for e in names[1:]:
        print(e.text)
    today = date.today()
    dt = today.strftime("%d-%m-%Y")
    wb = load_workbook(excel_f)
    sheet = wb[excel_sh]
    parse_column = config['column'][classroom]
    for name in names:
        for box in range(2, sheet.max_row+1):
            if name.text.lower() == (sheet.cell(row=box, column=column_index_from_string('A')).value.lower()):
                sheet[f'{parse_column}{box}'] = 'P'
    sheet[f'{parse_column}1'] = dt
    count = column_index_from_string(parse_column)
    count += 1
    increment_column = get_column_letter(count)
    config.set('column', classroom, increment_column)
    with open(config_file, 'w') as configfile:
        config.write(configfile)
    wb.save(excel_f)
    driver.quit()
    print(f'>>> No. of participants : {n-1}')
    notify('Attendance Taken Successfully!')
    print('\n>>> Attendance Taken Successfully!.........Thank You for using the service.\n')
    input('>>> Press any key to return to Main Menu')
    mainmenu()


if __name__ == '__main__':
    mainmenu()
